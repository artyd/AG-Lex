"""Contract ingestion: PDF/DOCX → Markdown → sections + token stats.

Phase 1.3. Used by `POST /api/upload`. The point of the conversion is two-fold:

1. Lawyers see the original; Claude works against the smaller Markdown.
2. Section splitting lets us later send Claude only the relevant clause instead
   of the whole contract — that's where the token savings actually come from.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Callable


# ---------------------------------------------------------------------------
# raw extraction (used for the token-savings comparison)
# ---------------------------------------------------------------------------

def pdf_raw_text(path: str | Path) -> str:
    import pymupdf
    doc = pymupdf.open(str(path))
    try:
        return "\n".join(page.get_text() for page in doc)
    finally:
        doc.close()


def docx_raw_text(path: str | Path) -> str:
    import mammoth
    with open(str(path), "rb") as f:
        return mammoth.extract_raw_text(f).value


def xlsx_raw_text(path: str | Path) -> str:
    """Flatten every non-empty cell across all sheets into newline-separated text."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c not in (None, "")]
                if cells:
                    lines.append("\t".join(cells))
        return "\n".join(lines)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# markdown conversion
# ---------------------------------------------------------------------------

def pdf_to_markdown(path: str | Path) -> str:
    """PyMuPDF4LLM produces a Markdown rendering with headings + tables."""
    import pymupdf4llm
    return pymupdf4llm.to_markdown(str(path))


def docx_to_markdown(path: str | Path) -> str:
    """Mammoth renders Word styling (headings, lists, tables) into Markdown."""
    import mammoth
    with open(str(path), "rb") as f:
        result = mammoth.convert_to_markdown(f)
    return result.value


def xlsx_to_markdown(path: str | Path) -> str:
    """Render every sheet as a Markdown table.

    Used for the procurement "Handover (Table 3)" which arrives as an Excel
    form: two-column N/Field/Value layouts and 3+ column tables both work.
    Empty rows are skipped; merged cells aren't expanded (openpyxl read-only
    flattens them to the top-left value, which matches how lawyers read the
    form anyway).
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for ws in wb.worksheets:
            rows = [
                ["" if c is None else str(c).strip() for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(cell for cell in r)]
            if not rows:
                continue
            width = max(len(r) for r in rows)
            rows = [r + [""] * (width - len(r)) for r in rows]
            if ws.title and ws.title.lower() not in {"sheet", "sheet1", "аркуш1"}:
                parts.append(f"## {ws.title}\n")
            header = rows[0]
            body = rows[1:]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join("---" for _ in header) + " |")
            for r in body:
                parts.append("| " + " | ".join(c.replace("|", "\\|") for c in r) + " |")
            parts.append("")
        return "\n".join(parts)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# HTML conversion (Phase 3.3) — used for the reconcile display so the FE
# can preserve the doc's look (tables, bilingual columns, lists) instead of
# flattening it via mammoth's markdown output. The Claude analysis still
# rides on the markdown above for token efficiency.
# ---------------------------------------------------------------------------

def pdf_to_html(path: str | Path) -> str:
    """Render every page as HTML and concatenate. pymupdf's HTML output
    preserves text positioning but flattens tables; good enough for display
    when the source is a born-digital PDF."""
    import pymupdf
    doc = pymupdf.open(str(path))
    try:
        parts = []
        for page in doc:
            parts.append(page.get_text("html"))
        return "\n".join(parts)
    finally:
        doc.close()


def docx_to_html(path: str | Path) -> str:
    """Mammoth's HTML output keeps Word tables, headings, lists, bold —
    everything we lose in the markdown round-trip. Inline styles are
    omitted; we restyle on the FE via .cmp-paper-src classes."""
    import mammoth
    with open(str(path), "rb") as f:
        result = mammoth.convert_to_html(f)
    return result.value


def xlsx_to_html(path: str | Path) -> str:
    """Render every sheet as an HTML table — mirrors xlsx_to_markdown but
    keeps real <table>/<tr>/<td> tags so the FE renderer doesn't have to
    parse pipe syntax for the handover view."""
    import openpyxl
    from html import escape as _esc
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for ws in wb.worksheets:
            rows = [
                ["" if c is None else str(c).strip() for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(cell for cell in r)]
            if not rows:
                continue
            if ws.title and ws.title.lower() not in {"sheet", "sheet1", "аркуш1"}:
                parts.append(f"<h3>{_esc(ws.title)}</h3>")
            width = max(len(r) for r in rows)
            rows = [r + [""] * (width - len(r)) for r in rows]
            parts.append("<table>")
            parts.append("<thead><tr>" + "".join(
                f"<th>{_esc(c)}</th>" for c in rows[0]
            ) + "</tr></thead>")
            parts.append("<tbody>")
            for r in rows[1:]:
                parts.append("<tr>" + "".join(
                    f"<td>{_esc(c)}</td>" for c in r
                ) + "</tr>")
            parts.append("</tbody></table>")
        return "\n".join(parts)
    finally:
        wb.close()


_CONVERTERS: dict[str, Callable[[str | Path], str]] = {
    ".pdf": pdf_to_markdown,
    ".docx": docx_to_markdown,
    ".xlsx": xlsx_to_markdown,
}

_HTML_CONVERTERS: dict[str, Callable[[str | Path], str]] = {
    ".pdf": pdf_to_html,
    ".docx": docx_to_html,
    ".xlsx": xlsx_to_html,
}

_RAW_EXTRACTORS: dict[str, Callable[[str | Path], str]] = {
    ".pdf": pdf_raw_text,
    ".docx": docx_raw_text,
    ".xlsx": xlsx_raw_text,
}


# ---------------------------------------------------------------------------
# Display PDF (Phase 4.x) — produce a single PDF the FE renders via PDF.js.
# Lawyers see the original 1-to-1 (with tables, columns, fonts) and the
# transparent highlight overlay paints over the actual words — no
# reconstruction in the browser. PDF inputs pass through unchanged; DOCX and
# XLSX go through headless LibreOffice (`soffice --headless --convert-to pdf`).
# ---------------------------------------------------------------------------

DISPLAY_PDF_EXTS: frozenset[str] = frozenset({".pdf", ".docx", ".xlsx"})


class DisplayPdfError(RuntimeError):
    """Raised when the display-PDF render fails.

    `kind` lets the caller decide how to surface the failure:
      missing    — soffice binary not on PATH / not executable
      crash      — soffice exited non-zero
      timeout    — render exceeded DISPLAY_PDF_TIMEOUT
      empty      — produced PDF is missing, not %PDF-, or absurdly small
      too_large  — produced PDF exceeds MAX_DISPLAY_PDF_BYTES
    """

    def __init__(self, kind: str, message: str = "") -> None:
        super().__init__(message or kind)
        self.kind = kind


def to_display_pdf(
    path: "str | Path",
    *,
    timeout: float | None = None,
    max_bytes: int | None = None,
    soffice_path: str | None = None,
) -> bytes:
    """Render any supported source (.pdf/.docx/.xlsx) to a display PDF.

    PDF inputs are returned as-is (no re-render — skips the ~2s soffice
    roundtrip when the user already gave us a PDF). DOCX/XLSX go through
    a headless `soffice --convert-to pdf` into a per-call mkdtemp so
    parallel calls never collide. The temp dir is cleaned in `finally`.

    Returns the raw PDF bytes. Raises `DisplayPdfError(kind=…)` on:
      - soffice binary missing / not executable
      - subprocess non-zero exit or timeout
      - empty / not-`%PDF-` output
      - output exceeding `max_bytes`

    Settings are read from `get_settings()` unless overridden via kwargs
    (useful for tests).
    """
    import os
    import shutil as _shutil
    import subprocess
    import sys
    import tempfile as _tempfile
    import time as _time
    from .config import get_settings

    settings = get_settings()
    timeout = settings.DISPLAY_PDF_TIMEOUT if timeout is None else timeout
    max_bytes = settings.MAX_DISPLAY_PDF_BYTES if max_bytes is None else max_bytes
    soffice = soffice_path if soffice_path is not None else settings.SOFFICE_PATH

    src = Path(path)
    suffix = src.suffix.lower()
    if suffix not in DISPLAY_PDF_EXTS:
        raise ValueError(
            f"Unsupported display-PDF source {suffix!r}. "
            f"Supported: {sorted(DISPLAY_PDF_EXTS)}"
        )

    if suffix == ".pdf":
        data = src.read_bytes()
        if not data.startswith(b"%PDF-"):
            raise DisplayPdfError("empty", f"input PDF is not a valid PDF: {src}")
        if len(data) > max_bytes:
            raise DisplayPdfError(
                "too_large",
                f"input PDF size {len(data)} exceeds cap {max_bytes}",
            )
        return data

    # DOCX/XLSX → soffice
    soffice_resolved = _shutil.which(soffice) if soffice else None
    # The fallback applies only to a bare command name like "soffice" (the
    # default). If the user gave us an explicit path (absolute or relative),
    # honor it and fail loud — silently rewriting a misconfigured absolute
    # path would hide bugs. `os.path.isabs` isn't reliable cross-platform
    # (Python 3.13 changed ntpath behavior for `/foo`), so check for path
    # separators directly.
    looks_like_bare_name = bool(soffice and "/" not in soffice and "\\" not in soffice)
    if not soffice_resolved and looks_like_bare_name:
        # which() failed for a bare command name. Common cause on prod: the
        # systemd unit's PATH excludes /usr/bin even though /usr/bin/soffice
        # exists (or `ProtectSystem=strict`/`PrivateUsers` hides it). Probe a
        # tight list of canonical install paths before declaring "missing" so
        # a vanilla install Just Works without forcing every host to set
        # SOFFICE_PATH. shutil.which() with an absolute path doesn't search
        # PATH — it only checks if the file exists and is executable.
        _FALLBACK_SOFFICE_PATHS = (
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        )
        for cand in _FALLBACK_SOFFICE_PATHS:
            hit = _shutil.which(cand)
            if hit:
                soffice_resolved = hit
                print(
                    f"[to_display_pdf] PATH lookup for {soffice!r} failed; "
                    f"using fallback {hit!r}. Set SOFFICE_PATH={hit} in the "
                    f"env to silence this lookup on subsequent calls.",
                    file=sys.stderr,
                    flush=True,
                )
                break
    if not soffice_resolved:
        raise DisplayPdfError(
            "missing",
            f"soffice binary not found (looked up {soffice!r}); "
            f"install LibreOffice on the host (apt-get install --no-install-recommends "
            f"libreoffice-core libreoffice-writer libreoffice-calc) "
            f"or set SOFFICE_PATH to an absolute path.",
        )

    outdir = _tempfile.mkdtemp(prefix="aglex_pdf_")
    # Per-call UserInstallation isolates the LibreOffice profile so two
    # concurrent conversions don't trip over each other's session locks.
    # Without this, parallel /api/reconcile uploads occasionally fail with
    # "another instance is already running" on the host.
    user_install = _tempfile.mkdtemp(prefix="aglex_lo_user_")
    try:
        cmd = [
            soffice_resolved,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--nolockcheck",
            f"-env:UserInstallation=file://{user_install}",
            "--convert-to", "pdf",
            "--outdir", outdir,
            str(src),
        ]
        # The soffice "binary" on Linux is a /bin/sh wrapper that shells out
        # to dirname/basename/sed/grep/uname to compute its own install dir.
        # When the parent process's PATH excludes /usr/bin (e.g. a systemd
        # unit with Environment="PATH=/app/venv/bin"), the wrapper can't
        # find coreutils and exits 127 with cryptic "exec: /<workdir>/oosplash:
        # not found" — because dirname failed, so it fell back to $PWD. Force
        # /usr/bin and /bin onto the child's PATH so the wrapper Just Works
        # regardless of how the host configured the parent's PATH.
        child_env = dict(os.environ)
        _SYSTEM_BIN_DIRS = ("/usr/bin", "/bin", "/usr/sbin", "/sbin")
        path_segments = (child_env.get("PATH") or "").split(os.pathsep)
        path_segments = [p for p in path_segments if p]
        for d in reversed(_SYSTEM_BIN_DIRS):
            if d not in path_segments:
                path_segments.insert(0, d)
        child_env["PATH"] = os.pathsep.join(path_segments)
        t0 = _time.perf_counter()
        try:
            proc = subprocess.run(
                cmd,
                timeout=timeout,
                capture_output=True,
                check=False,
                env=child_env,
            )
        except subprocess.TimeoutExpired as e:
            raise DisplayPdfError(
                "timeout",
                f"soffice exceeded {timeout}s rendering {src.name}",
            ) from e
        elapsed = _time.perf_counter() - t0

        # Always echo what soffice said — even on success — so the systemd
        # journal carries enough context when a deployment regresses.
        stdout_tail = (proc.stdout or b"")[-512:].decode("utf-8", errors="replace")
        stderr_tail = (proc.stderr or b"")[-512:].decode("utf-8", errors="replace")
        if proc.returncode != 0:
            print(
                f"[to_display_pdf] soffice exit {proc.returncode} for {src.name} "
                f"after {elapsed:.2f}s\n  cmd: {cmd!r}\n  stdout: {stdout_tail!r}\n  stderr: {stderr_tail!r}",
                file=sys.stderr,
                flush=True,
            )
            # Surface what soffice actually complained about so the UI banner
            # can show the cause without sending the user to journalctl. Cap
            # the tail aggressively (exit 127 → "oosplash: not found" type
            # messages; we don't need 512 chars of XML stack traces in the UI).
            tail = (stderr_tail or stdout_tail or "").strip()
            if len(tail) > 240:
                tail = "…" + tail[-240:]
            detail = f"soffice exit {proc.returncode} rendering {src.name}"
            if tail:
                detail = f"{detail}: {tail}"
            raise DisplayPdfError("crash", detail)

        produced = Path(outdir) / (src.stem + ".pdf")
        if not produced.is_file():
            # soffice sometimes renames the output when the stem contains
            # a non-ASCII char — fall back to "whatever .pdf we got".
            pdfs = list(Path(outdir).glob("*.pdf"))
            if pdfs:
                produced = pdfs[0]
        if not produced.is_file():
            print(
                f"[to_display_pdf] soffice produced no PDF for {src.name}\n"
                f"  outdir contents: {list(Path(outdir).iterdir())!r}\n"
                f"  stdout: {stdout_tail!r}\n  stderr: {stderr_tail!r}",
                file=sys.stderr,
                flush=True,
            )
            raise DisplayPdfError(
                "empty",
                f"soffice produced no output PDF for {src.name}",
            )
        data = produced.read_bytes()
        # 100 bytes is a generous "must at least look like a PDF" floor —
        # a single-cell xlsx render can be remarkably small. 256 was rejecting
        # legitimate handover files.
        if len(data) < 100 or not data.startswith(b"%PDF-"):
            print(
                f"[to_display_pdf] soffice output for {src.name} looks invalid "
                f"({len(data)} bytes, head: {data[:32]!r})",
                file=sys.stderr,
                flush=True,
            )
            raise DisplayPdfError(
                "empty",
                f"soffice output for {src.name} is not a valid PDF ({len(data)} bytes)",
            )
        if len(data) > max_bytes:
            raise DisplayPdfError(
                "too_large",
                f"display PDF for {src.name} is {len(data)} bytes (cap {max_bytes})",
            )

        # Lightweight observability — the systemd journal already collects stderr.
        print(
            f"[to_display_pdf] OK {src.name}: {len(data)} bytes in {elapsed:.2f}s",
            file=sys.stderr,
            flush=True,
        )
        return data
    finally:
        _shutil.rmtree(outdir, ignore_errors=True)
        _shutil.rmtree(user_install, ignore_errors=True)


def detect_type_and_convert(path: str | Path) -> str:
    suffix = Path(path).suffix.lower()
    fn = _CONVERTERS.get(suffix)
    if fn is None:
        raise ValueError(f"Unsupported file type: {suffix!r}. Supported: .pdf, .docx, .xlsx")
    return fn(path)


def detect_type_and_convert_html(path: str | Path) -> str:
    """Display-side renderer: returns HTML preserving the doc's structure
    (tables, headings, lists). Used by /api/reconcile so the FE shows the
    original look instead of mammoth's flattened markdown."""
    suffix = Path(path).suffix.lower()
    fn = _HTML_CONVERTERS.get(suffix)
    if fn is None:
        raise ValueError(f"Unsupported file type: {suffix!r}. Supported: .pdf, .docx, .xlsx")
    return fn(path)


def detect_type_and_extract_raw(path: str | Path) -> str:
    suffix = Path(path).suffix.lower()
    fn = _RAW_EXTRACTORS.get(suffix)
    if fn is None:
        raise ValueError(f"Unsupported file type: {suffix!r}. Supported: .pdf, .docx, .xlsx")
    return fn(path)


# ---------------------------------------------------------------------------
# section splitting
# ---------------------------------------------------------------------------

# Contract section markers. Anchored to line-start so inline references like
# "як зазначено у Статті 5" don't trigger a split. Leading markdown decorations
# (`#`, `**`, `>`, list bullets) are tolerated. The bare-number rule
# `\d+\.(?=\s+\S)` only fires when followed by a non-empty heading word — that
# trims most numbered-list false positives.
_SECTION_RE = re.compile(
    r"^"
    r"[ \t]*(?:[#>\-\*_]+[ \t]*)*"
    r"(?P<number>"
    # Headings: capture digits/Roman numerals but stop at the separating dot so
    # "Стаття 5." doesn't include the trailing dot in the number.
    r"(?:Розділ|Стаття)\s+[^\s\n.]+"
    r"|п\.\s*\d+(?:\.\d+)*"
    r"|\d+(?:\.\d+)+"
    # Bare "1." / "2." numbered points — keep the dot so the label stays
    # distinguishable from a plain integer.
    r"|\d+\.(?=\s+\S)"
    r")"
    r"[ \t\.]*"
    r"(?P<title>[^\n]*)$",
    re.MULTILINE | re.IGNORECASE,
)


def _clean_title(raw: str) -> str | None:
    cleaned = re.sub(r"[\*_`#>]+", " ", raw).strip(" .-–—")
    return cleaned or None


def split_into_sections(markdown: str) -> list[dict]:
    """Slice contract Markdown into `{number, title, text}` sections.

    Splitting is best-effort by design (the Phase 1.3 doc explicitly says even
    rough cuts already give token savings). Empty bodies and a noisy preamble
    before the first marker are kept as a section with `number=None` only when
    they contain visible text.
    """
    matches = list(_SECTION_RE.finditer(markdown))
    if not matches:
        return [{"number": None, "title": None, "text": markdown.strip()}] if markdown.strip() else []

    sections: list[dict] = []

    head = markdown[: matches[0].start()].strip()
    if head:
        sections.append({"number": None, "title": None, "text": head})

    for i, m in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(markdown)
        body = markdown[m.end():end].strip()
        sections.append(
            {
                "number": m.group("number").strip(),
                "title": _clean_title(m.group("title") or ""),
                "text": body,
            }
        )
    return sections


# ---------------------------------------------------------------------------
# token estimation
# ---------------------------------------------------------------------------

def estimate_tokens(text: str | None) -> int:
    """Rough heuristic: ~4 characters per token.

    Holds well enough for both English and Ukrainian Cyrillic. Off by ~20% in
    either direction depending on the model, but adequate for the "Markdown vs
    raw" comparison we surface in `/api/upload`. Swap in tiktoken/anthropic
    tokenizer later if exact counts matter.
    """
    if not text:
        return 0
    return max(1, len(text) // 4)


def token_savings(raw: str, markdown: str) -> dict:
    raw_t = estimate_tokens(raw)
    md_t = estimate_tokens(markdown)
    pct = round((1 - md_t / raw_t) * 100, 1) if raw_t else 0.0
    return {"raw_tokens": raw_t, "markdown_tokens": md_t, "savings_pct": pct}
